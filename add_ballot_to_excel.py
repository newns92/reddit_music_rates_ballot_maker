## TODO: Function definition and comments
import pandas as pd
import argparse
import os
from openpyxl import load_workbook


def find_cells_by_value(filename, sheet_name, search_value_bonus, search_value_end):
    ## Load the workbook (use read_only=True for performance if only reading)
    wb = load_workbook(filename, read_only=True)
    ws = wb[sheet_name]

    found_end_cells = []
    found_album_cells = []
    
    ## Iterate over all rows in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            ## Check if the cell's value is "END"
            if cell.value == search_value_end:
                found_end_cells.append(cell)
            ## If not, check for "BONUS TRACKS" and save that cell, if needed
            elif cell.value == search_value_bonus:
                bonus_cell = cell
            elif cell.value[0:6] == 'Album:':
                found_album_cells.append(cell)

    if found_end_cells:
        print(f"Found '{search_value_end}' in the following cells:")
        for cell in found_end_cells:
            print(f"- END Cell coordinate: {cell.coordinate}, Row: {cell.row}, Column: {cell.column}")
            # Example: print the value from the second column (index 1 in 0-based list) of the same row
            # Note: with read_only=True and iter_rows, accessing cells by index is easier
            # Example to get another value if needed can be handled in a separate iteration if using read_only
            # for full functionality or use the below approach if not using read_only
            # print(f"  Value in column 2 of this row: {ws.cell(row=cell.row, column=2).value}") # This requires not using read_only=True
    else:
        print(f"Value '{search_value_end}' not found in the sheet.")

    if bonus_cell:
        print(f"Found '{search_value_bonus}' in the following cells:")
        print(f"- BONUS TRACKS Cell coordinate: {bonus_cell.coordinate}, Row: {bonus_cell.row}, Column: {bonus_cell.column}")
    else:
        print(f"Value '{search_value_bonus}' not found in the sheet.")

    if found_album_cells:
        print("Found albums in the following cells:")
        for cell in found_album_cells:
            print(f"- Album Cell coordinate: {cell.coordinate}, Row: {cell.row}, Column: {cell.column}")
    else:
        print('No album cells found in the sheet.')
        
    wb.close()

    ## Get the BONUS TRACKS cell row number, if found
    final_cells = {}
    final_cells['bonus'] = bonus_cell if bonus_cell else None

    ## Then, get the max row number from the found END cell(s) to get to bottom of ballot
    ##  - i.e., After a bonus bonus rate, if exists
    final_cells['end'] = found_end_cells[-1] if found_end_cells else None

    if found_album_cells:
        for i in range(len(found_album_cells)):
            final_cells[f'album_{i+1}'] = found_album_cells[i]

    return final_cells


def create_function_strings(final_cells):
    ## TODO: Edit in case it's a grab bag, so there are no Album ranges
    ## Extract and sort album cells by row number
    ## This ensures the order of album cell coordinates is correct,
    ##  regardless of dictionary insertion order 
    album_cells = sorted(
        ## Iterable
        [(key, val) for key, val in final_cells.items() if key.startswith('album')],
        ## Sort by row number of the cell value (an openpyxl 'cell' object)
        key = lambda x: x[1].row
    )
    # print(album_cells)

    bonus_row = final_cells['bonus'].row
    album_ranges = []

    ## For each key-cell pair in the list...
    for i, (key, cell) in enumerate(album_cells):
        # print(i)
        ## Skip "Album:" row
        range_start = cell.row + 1  
        
        ## If we still have albums to go through (until we reach end of len())...
        if i + 1 < len(album_cells):
            ## Album range ends one row above the *next* ([i + 1]) album
            range_end = album_cells[i + 1][1].row - 1
        ## Once our index i = length of album_cells list (we are on the final album)...
        else:
            ## Album range ends one row above 'BONUS TRACKS" cell
            range_end = bonus_row - 1

        album_ranges.append(f'B{range_start}:B{range_end}')

    joined_ranges = ','.join(album_ranges)
    joined_full_range = f'{joined_ranges[0:2]}:{joined_ranges[-3:]}'

    ## Prefix formula string with _xlfn. to avoid putting @ or {} in function cell value
    average_excel_function = f'=_xlfn.AVERAGE({joined_ranges})'
    median_excel_function = f'=_xlfn.MEDIAN({joined_ranges})'
    tens_excel_function = f'=_xlfn.COUNTIF({joined_full_range},">=10")'
    sub_fives_excel_function = f'=_xlfn.COUNTIF({joined_full_range},"<5")'
    mode_excel_function = f'=_xlfn.MODE.SNGL({joined_ranges})'
    std_dev_excel_function = f'=_xlfn.STDEV.S({joined_ranges})'

    functions = {}
    functions['Average'] = average_excel_function
    functions['Median'] = median_excel_function
    functions['Tens'] = tens_excel_function
    functions['SubFives'] = sub_fives_excel_function
    functions['Mode'] = mode_excel_function
    functions['StdDev'] = std_dev_excel_function

    return functions


def insert_stats_cells(file, sheet_name, final_cells, functions):
    """
    EXPLAIN

    Args:
        filepath (str): Path to the .xlsx file.
        search_value: The value to search for in the sheet.
        values_to_insert (list): A list of values/formulas to insert 
                                  horizontally (left to right) from the offset cell.
        sheet_name (str): Optional sheet name. Defaults to the active sheet.
    """
    wb = load_workbook(file)
    ws = wb[sheet_name]

    end_row, end_col = None, None

    ## Check if the 'BONUS TRACKS' cell was found and get coordinates if so
    end_cell = final_cells['end']
    if end_cell:
        end_row = end_cell.row
        end_col = end_cell.column
    
    # print(end_coord, end_row, end_col)

    ## Move one row down, one column to the right
    insert_row = end_row + 1
    # insert_col = end_col + 1

    ## Write each value/formula into the correct cells
    for i, item in enumerate(functions.items()):
        ## Break out key and value from dictionary item tuple
        key, value = item
        # print(f'Inserting "{key}" at row {insert_row + i}, column {insert_col}')
        # print(f'Inserting "{value}" at row {insert_row + i}, column {insert_col + 1}')
        formula_name_cell = f'B{str(insert_row + i)}'
        formula_cell = f'C{str(insert_row + i)}'
        ## Write to the above cells
        ws[formula_name_cell] = key
        ws[formula_cell] = value

    wb.save(file)
    wb.close()


def create_sheet(file, new_sheet_name):
    ## 1. Load the ballot into a DataFrame
    df = pd.read_csv('indieheads_ult_26.txt', 
                     sep='\r', 
                     header=None
                )

    ## 2. Use Pandas ExcelWriter to append the DataFrame as a new sheet
    try:
        with pd.ExcelWriter(file,
                            mode='a',  ## Open in append mode
                            engine='openpyxl', 
                            if_sheet_exists='replace'
                        ) as writer:
                df.to_excel(writer,
                            sheet_name=new_sheet_name,
                            index=False, 
                            header=False
                        )
                
        print(f'DataFrame successfully written to sheet {new_sheet_name} in {file}')

    except FileNotFoundError:
        print(f"Error: The file '{file}' was not found. Please ensure the file exists.")
    except Exception as e:
        print(f"An error occurred: {e}")

    ## 3. Find where alsbums end and ballot ENDs in order to start the statisics portion
    final_cells = find_cells_by_value(file, new_sheet_name, 'BONUS TRACKS', 'END')
    # print(final_cells)

    ## 4. Create the Excel formula strings for the statistics to be added
    excel_functions = create_function_strings(final_cells)
    print(excel_functions)

    ## 5. Insert the statistics formulas into the appropriate cells
    insert_stats_cells(file, new_sheet_name, final_cells, excel_functions)

    
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('filepath', help='Rates Excel file path')
    parser.add_argument('new_sheet_name', help='Name for new sheet for the ballot')
    args = parser.parse_args()
    
    create_sheet(args.filepath, args.new_sheet_name)
